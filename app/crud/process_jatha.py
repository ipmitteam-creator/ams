import os
import re
import pandas as pd
from openpyxl import load_workbook

# === CONFIG ===
CURRENT_FOLDER = os.path.dirname(os.path.abspath(__file__))  # same folder as script
TEMPLATE_FILE = os.path.join(CURRENT_FOLDER, "Jatha_Attendance_Bhati_Beas May 2025.xlsx")
OUTPUT_FILE = os.path.join(CURRENT_FOLDER, "Final_Jatha_Attendance.xlsx")
SHEET_NAME = "BHATI & BEAS"  # From your template

# Regex to extract date from filename (format: DD.MM.YYYY)
date_pattern = re.compile(r"(\d{2}\.\d{2}\.\d{4})")

# Collect all extracted rows
all_data = []

# Loop through all input files in same folder
for file in os.listdir(CURRENT_FOLDER):
    if file.endswith(".xlsx") and not file.startswith("~") and "Jatha_Attendance" not in file:
        file_path = os.path.join(CURRENT_FOLDER, file)

        # Extract date from filename
        match = date_pattern.search(file)
        date_val = match.group(1) if match else ""

        # Read input Excel (row 12 = index 11, so header=10)
        df = pd.read_excel(file_path, header=10)

        # Find Badge and Name columns dynamically
        badge_col = next((c for c in df.columns if "Badge" in str(c)), None)
        name_col = next((c for c in df.columns if "Name" in str(c)), None)

        if badge_col and name_col:
            filtered = df[df[badge_col].astype(str).str.startswith("GB")]
            for _, row in filtered.iterrows():
                all_data.append([row[badge_col], row[name_col], date_val])

# === Write into template ===
wb = load_workbook(TEMPLATE_FILE)
ws = wb[SHEET_NAME]

start_row = 8
for i, (badge, name, date_val) in enumerate(all_data, start=start_row):
    ws[f"A{i}"] = badge     # Badge Number
    ws[f"D{i}"] = name      # Name of Sewadar
    ws[f"E{i}"] = date_val  # Date from filename

wb.save(OUTPUT_FILE)
print(f"✅ Final file saved as {OUTPUT_FILE}")
