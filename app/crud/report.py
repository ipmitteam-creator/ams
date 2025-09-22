import psycopg2
from datetime import date, timedelta
from io import BytesIO
from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

router = APIRouter()

DB_CONFIG = {
    "dbname": "ams",
    "user": "neondb_owner",
    "password": "npg_igo8fBOT3MtP",
    "host": "ep-orange-fog-a1qfrxr9-pooler.ap-southeast-1.aws.neon.tech",
    "port": "5432",
    "sslmode": "require"
}

def get_db_connection():
    return psycopg2.connect(**DB_CONFIG)

# === JATHA REPORT ===
@router.get("/report/jatha")
def get_jatha_report(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(None, description="End date (YYYY-MM-DD). Defaults to start_date if not provided"),
    department: str = Query(None, description="Working department filter (optional)"),
    badge_no: str = Query(None, description="Badge number filter (optional)"),
    format: str = Query("json", description="Response format: json or xlsx")
):
    valid_jatha_types = ["Beas", "Bhati", "Others"]

    if not end_date:
        end_date = start_date
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="end_date cannot be before start_date.")

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT s.badge_no, s.name, a.attendance_date, a.attendance_type
        FROM attendance a
        JOIN sewadar s ON a.sewadar_id = s.sewadar_id
        JOIN department d ON s.current_department_id = d.department_id
        WHERE a.attendance_date BETWEEN %s AND %s
        AND a.attendance_type = ANY(%s)
    """
    params = [start_date, end_date, valid_jatha_types]
    if department:
        query += " AND d.department_name = %s"
        params.append(department)
    if badge_no:
        query += " AND s.badge_no = %s"
        params.append(badge_no)

    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        if format.lower() == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "NoData"
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            filename = f"Jatha_Report_{start_date.strftime('%d-%m-%Y')}_{end_date.strftime('%d-%m-%Y')}.xlsx"
            return StreamingResponse(stream,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'}
            )
        else:
            return {
                "start_date": start_date.strftime("%d/%m/%Y"),
                "end_date": end_date.strftime("%d/%m/%Y"),
                "sheets": {t: [] for t in valid_jatha_types}
            }

    grouped = {t: [] for t in valid_jatha_types}
    for badge, name, attendance_date, a_type in rows:
        grouped[a_type].append({
            "badge_no": badge,
            "duty_type": "J",
            "date_of_seva": attendance_date.strftime("%d/%m/%Y"),
            "name": name
        })

    for a_type in grouped:
        grouped[a_type].sort(key=lambda x: (x["date_of_seva"], x["badge_no"]))

    if format.lower() == "xlsx":
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, data in grouped.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(["Badge Number", "Duty Type", "Date of Seva DD/MM/YYYY", "Name of Sewadar"])
            for row in data:
                ws.append([row["badge_no"], row["duty_type"], row["date_of_seva"], row["name"]])
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"Jatha_Report_{start_date.strftime('%d-%m-%Y')}_{end_date.strftime('%d-%m-%Y')}.xlsx"
        return StreamingResponse(stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'}
        )

    return {
        "start_date": start_date.strftime("%d/%m/%Y"),
        "end_date": end_date.strftime("%d/%m/%Y"),
        "sheets": grouped
    }

# === ATTENDANCE REPORT ===
@router.get("/report")
def get_attendance_report(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(None, description="End date (YYYY-MM-DD). Defaults to start_date if not provided"),
    attendance_type: str = Query(None, description="Attendance type filter (optional)"),
    department: str = Query(None, description="Working department filter (optional)"),
    badge_no: str = Query(None, description="Badge number filter (optional)")
):
    if not end_date:
        end_date = start_date

    conn = get_db_connection()
    cur = conn.cursor()

    base_query = """
        SELECT s.sewadar_id, s.badge_no, s.name, d.department_name AS current_department_name
        FROM sewadar s
        JOIN department d ON s.current_department_id = d.department_id
    """
    filters = []
    params = []
    if department:
        filters.append("d.department_name = %s")
        params.append(department)
    if badge_no:
        filters.append("s.badge_no = %s")
        params.append(badge_no)
    if filters:
        base_query += " WHERE " + " AND ".join(filters)

    cur.execute(base_query, tuple(params))
    sewadaars = cur.fetchall()
    if not sewadaars:
        conn.close()
        raise HTTPException(status_code=404, detail="No sewadaars found for given filters.")

    attendance_query = """
        SELECT sewadar_id, attendance_date, attendance_type
        FROM attendance
        WHERE attendance_date BETWEEN %s AND %s
    """
    attendance_params = [start_date, end_date]
    if attendance_type:
        attendance_query += " AND attendance_type = %s"
        attendance_params.append(attendance_type)

    cur.execute(attendance_query, tuple(attendance_params))
    attendance_records = cur.fetchall()
    conn.close()

    attendance_lookup = {(r[0], r[1]): r[2] for r in attendance_records}

    report = []
    day_count = (end_date - start_date).days + 1
    for offset in range(day_count):
        day = start_date + timedelta(days=offset)
        for sewadar_id, badge, name, dept_name in sewadaars:
            key = (sewadar_id, day)
            if key in attendance_lookup:
                a_type = attendance_lookup[key]
                status = "Jatha" if a_type in ["Bhati", "Beas", "Others"] else "Present"
            else:
                status = "Absent"
            report.append({
                "date": str(day),
                "name": name,
                "badge_no": badge,
                "current_department_name": dept_name,
                "status": status
            })

    return {
        "start_date": str(start_date),
        "end_date": str(end_date),
        "attendance_type_filter": attendance_type or "All",
        "department_filter": department or "All",
        "badge_filter": badge_no or "All",
        "report": report
    }
